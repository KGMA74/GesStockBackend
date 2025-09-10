from .serializers import (
    CustomUserSerializer, CreateUserSerializer,
    ProductSerializer, StockEntrySerializer, StockExitSerializer,
    StockEntryFormSerializer, StockExitFormSerializer,
    SupplierSerializer, WarehouseSerializer, CustomerSerializer, AccountSerializer, InvoiceSerializer,
    FinancialTransactionSerializer, StockTransferSerializer, StockTransferFormSerializer,
    DebtPaymentSerializer
)
from django.contrib.auth import login, user_logged_in
from .models import User, Product, StockEntry, StockExit, StockEntryItem, StockExitItem, Supplier, Warehouse, Customer, ProductStock, Account, Invoice, FinancialTransaction, StockTransfer, StockTransferItem
from django.conf import settings 
from djoser.views import UserViewSet
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db.models import Q, Sum, Count, F
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework.exceptions import ValidationError
from rest_framework import generics, status, viewsets
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny
from rest_framework.decorators import api_view, action, permission_classes
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from djoser.social.views import ProviderAuthView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.exceptions import TokenError
from .authentication import CustomAuthenticationBackend
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML
import io


# from authentication.permissions import CanApproveLevel1, CanApproveLevel2, CanApproveLevel3
from rest_framework_simplejwt.views import (
    TokenObtainPairView,
    TokenRefreshView,
    TokenVerifyView
)
from .utils import set_auth_cookie
from .models import Store
# Permission
from django.contrib.auth import user_logged_in
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.core.exceptions import ObjectDoesNotExist
import logging

logger = logging.getLogger(__name__)

class StoreContextMixin:
    """
    Mixin pour gérer le contexte store dans les vues.
    Assure que l'utilisateur a un store valide et active.
    """
    
    @property
    def store_id(self):
        """Retourne l'ID du store de l'utilisateur connecté."""
        store = self.store
        return store.id if store else None
    
    @property
    def store(self):
        """Retourne le store de l'utilisateur connecté."""
        if not hasattr(self.request, 'user') or not self.request.user.is_authenticated:
            return None
        return getattr(self.request.user, 'store', None)

    def dispatch(self, request, *args, **kwargs):
        """
        Vérifie que l'utilisateur a un store valide avant de traiter la requête.
        """
        try:
            store = self.store
            if not store:
                logger.warning(f"Tentative d'accès sans store valide par l'utilisateur {request.user.id}")
                return Response(
                    {
                        'error': 'Marchand non autorisé ou non trouvé',
                        'detail': 'Votre compte doit être associé à un marchand valide.'
                    }, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            if not store.is_active:
                logger.warning(f"Tentative d'accès avec store inactif {store.id} par l'utilisateur {request.user.id}")
                return Response(
                    {
                        'error': 'Marchand désactivé',
                        'detail': 'Votre marchand est actuellement désactivé.'
                    },
                    status=status.HTTP_403_FORBIDDEN
                )
                
        except Exception as e:
            logger.error(f"Erreur lors de la vérification du store: {str(e)}")
            return Response(
                {'error': 'Erreur interne du serveur'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
        return super().dispatch(request, *args, **kwargs)
    
    def get_store_queryset(self, queryset):
        """
        Filtre le queryset pour ne retourner que les objets du store actuel.
        """
        store = self.store
        if not store:
            return queryset.none()
            
        # Vérifier si le modèle a un champ store
        if hasattr(queryset.model, 'store'):
            return queryset.filter(store=store)
        
        # Si pas de champ store, retourner le queryset complet
        # (à adapter selon la logique métier)
        return queryset
    

def test(request):
    from django.shortcuts import render
    return render(request, 'accounts/test.html')



class CustomProviderAuthView(ProviderAuthView):
    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)

        if response.status_code == 201 and response.data:
            access_token = response.data.get('access')
            refresh_token = response.data.get('refresh')
            if access_token:
                set_auth_cookie(response, 'access', access_token, settings.AUTH_COOKIE_ACCESS_MAX_AGE)
            if refresh_token:
                set_auth_cookie(response, 'refresh', refresh_token, settings.AUTH_COOKIE_REFRESH_MAX_AGE)
        return response
    
#customisation de la class TokenObtainPairView pour que les tokens passes par les cookies et non les headers donc plus securise
class CustomTokenObtainPairView(TokenObtainPairView, StoreContextMixin):
    def post(self, request, *args, **kwargs):
        _response = super().post(request, *args, **kwargs)
        response = Response({}, status=status.HTTP_200_OK)
        
        if _response.status_code == 200 and _response.data:
            access_token = _response.data.get('access')
            refresh_token = _response.data.get('refresh')

            if access_token:
                set_auth_cookie(response, 'access', access_token, settings.AUTH_COOKIE_ACCESS_MAX_AGE)
                
            if refresh_token:
                set_auth_cookie(response, 'refresh', refresh_token, settings.AUTH_COOKIE_REFRESH_MAX_AGE)
    
            try:

                user = _response.data.get('user')
                
                # Si l'utilisateur est staff, créer une session Django
                # TODO:  remove this later
                if user.is_staff:
                    login(request, user, backend='api.authentication.CustomAuthenticationBackend')
                else:
                    pass
                    
                user_serializer = CustomUserSerializer(user, context = {'request': request})
                response.data = user_serializer.data
            
            except User.DoesNotExist:
                pass
            
            return response

        return _response


class CustomTokenRefreshView(TokenRefreshView):
    def post(self, request, *args, **kwargs):
        refresh_token = request.COOKIES.get('refresh')
        
        if refresh_token:
            mutable_data = request.data.copy()
            mutable_data['refresh'] = refresh_token
            request._full_data = mutable_data
            
            
        _response = super().post(request, *args, **kwargs)
        response = Response(status=status.HTTP_200_OK)
        print(_response)
        if _response.status_code == 200 and _response.data:
            access_token = _response.data.get('access')
            refresh_token = _response.data.get('refresh')
            
            # Update the tokens
            if access_token:
                set_auth_cookie(response, 'access', access_token, settings.AUTH_COOKIE_ACCESS_MAX_AGE)
                
            if refresh_token:
                set_auth_cookie(response, 'refresh', refresh_token, settings.AUTH_COOKIE_REFRESH_MAX_AGE)
            
            return response
        
        return _response

class CustomTokenVerifyView(TokenVerifyView):
    def post(self, request, *args, **kwargs):
        if not request.data.get('token') and request.COOKIES.get('access'):
            data = request.data.copy()
            
            data['token'] = request.COOKIES.get('access')
            request.data = data
            
        return super().post(request, *args, **kwargs)
    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def LogoutView(request):
    try:
        refresh_token = request.COOKIES.get('refresh')
        response = Response({"detail": "Déconnexion réussie."}, status=status.HTTP_205_RESET_CONTENT)
        
        if refresh_token:
            token = RefreshToken(refresh_token)
            token.blacklist()
            
        # Supprime les cookies
        set_auth_cookie(response, 'access', '', max_age=0)
        set_auth_cookie(response, 'refresh', '', max_age=0)
        
        request.session.flush()
        return response
            
    except TokenError as e:
        return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({"detail": "Une erreur s'est produite lors de la déconnexion"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CustomUserViewSet(UserViewSet, StoreContextMixin):
    def get_serializer_class(self):
        if self.action == 'create':
            return CreateUserSerializer
        return super().get_serializer_class()
    
    def get_queryset(self):
        """Retourne le queryset ordonné des User du store courant."""
        return self.get_store_queryset(User.objects.all().order_by('-created_at', 'fullname'))
    
    def get_permissions(self):
        if self.action == 'list':
            return [IsAuthenticated()]
        if self.action == 'retrieve':
            return [IsAuthenticated()]
        if self.action == 'create':
            return [IsAuthenticated()] #ajout de permission personnaliser pour lajout
        return super().get_permissions()
    
    def create(self, request, *args, **kwargs):        
        serializer = self.get_serializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        
        try:
            with transaction.atomic():
                user = serializer.save()

                if user.is_superuser and not request.user.is_superuser:
                    raise ValidationError("You cannot create a superuser.")
        except:
            return Response({"error": "Un utilisateur avec ces identifiants existe deja"}, status=status.HTTP_400_BAD_REQUEST)
        return Response({
            "detail": "Utilisateur créé avec succès",
            "user_id": user.id
        }, status=status.HTTP_201_CREATED)
        
        
    @action(["post"], detail=False)
    def reset_password(self, request, *args, **kwargs):
        from djoser.compat import get_user_email
        from djoser.conf import settings
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.get_user()

        if user:
            context = {"user": user, "store_code": user.store.store_code}
            to = [get_user_email(user)]
            settings.EMAIL.password_reset(self.request, context).send(to)

        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(methods=['post'], detail=True, url_path='toggle-activation-status')
    def toggle_activation_status(self, request, id=None):
        user = self.get_object()
        user.is_active = not user.is_active
        user.save(update_fields=['is_active'])
        return Response(status=status.HTTP_200_OK)
    
    @action(methods=['post'], detail=True, url_path='grant-permission')
    def grant_permission(self, request, id=None):
        user = self.get_object()
        user.grant_permission(request.data.get('permission_code'), granted_by=request.user)
        return Response(status=status.HTTP_200_OK)
    
    @action(methods=['post'], detail=True, url_path='revoke-permission')
    def revoke_permission(self, request, id=None):
        user = self.get_object()
        user.revoke_permission(request.data.get('permission_code'))
        return Response(status=status.HTTP_200_OK)
    
class testViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = CustomUserSerializer


# ViewSets pour la gestion de stock
class ProductViewSet(viewsets.ModelViewSet, StoreContextMixin):
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]  # Temporairement sans DjangoFilterBackend
    search_fields = ['name', 'reference', 'description']
    # filterset_fields = []  # Désactivé temporairement
    ordering_fields = ['name', 'created_at', 'reference']
    ordering = ['-created_at']
    
    def get_queryset(self):
        queryset = Product.objects.all().order_by('-created_at')
        return self.get_store_queryset(queryset)
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['store'] = self.store
        return context
    
    def perform_create(self, serializer):
        print(self.store)
        serializer.save(store=self.store)
    
    @action(detail=False, methods=['get'], url_path='low-stock')
    def low_stock(self, request):
        """Retourne les produits avec un stock faible"""
        queryset = self.get_queryset()
        low_stock_products = []
        
        for product in queryset:
            # Calculer le stock actuel de manière similaire à stock_stats
            entries = StockEntryItem.objects.filter(product=product).aggregate(
                total=Sum('quantity'))['total'] or 0
            exits = StockExitItem.objects.filter(product=product).aggregate(
                total=Sum('quantity'))['total'] or 0
            current_stock = entries - exits
            
            if current_stock <= product.min_stock_alert:
                low_stock_products.append(product)
        
        serializer = self.get_serializer(low_stock_products, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='search')
    def search(self, request):
        """Recherche rapide pour l'autocomplétion"""
        search_query = request.query_params.get('search', '')
        limit = int(request.query_params.get('limit', 20))
        
        if not search_query:
            return Response([])
        
        # Obtenir le queryset de base sans l'ordre par défaut
        base_queryset = Product.objects.all()
        base_queryset = self.get_store_queryset(base_queryset)
        
        # Recherche avec ordre de pertinence
        from django.db.models import Case, When, Value, IntegerField
        
        queryset = base_queryset.filter(
            Q(name__icontains=search_query) |
            Q(reference__icontains=search_query) |
            Q(description__icontains=search_query)
        ).annotate(
            # Priorité de recherche : référence exacte > référence partielle > nom partielle
            search_priority=Case(
                When(reference__iexact=search_query, then=Value(1)),
                When(reference__icontains=search_query, then=Value(2)),
                When(name__icontains=search_query, then=Value(3)),
                default=Value(4),
                output_field=IntegerField()
            )
        ).order_by('search_priority', 'reference', 'name')[:limit]
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class CustomerViewSet(viewsets.ModelViewSet, StoreContextMixin):
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'phone', 'email', 'address']
    ordering_fields = ['name', 'created_at']
    ordering = ['-created_at']
    
    def get_queryset(self):
        queryset = Customer.objects.all().order_by('-created_at')
        return self.get_store_queryset(queryset)
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['store'] = self.store
        return context
    
    def perform_create(self, serializer):
        serializer.save(store=self.store)
    
    @action(detail=False, methods=['get'], url_path='search')
    def search(self, request):
        """Recherche rapide pour l'autocomplétion"""
        search_query = request.query_params.get('search', '')
        limit = int(request.query_params.get('limit', 20))
        
        if not search_query:
            return Response([])
        
        queryset = self.get_queryset().filter(
            Q(name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(address__icontains=search_query)
        )[:limit]
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], url_path='pay-debt')
    def pay_debt(self, request, pk=None):
        """Enregistrer un remboursement de dette client"""
        from .serializers import DebtPaymentSerializer
        
        try:
            customer = self.get_object()
            
            if customer.debt <= 0:
                return Response(
                    {'error': 'Ce client n\'a pas de dette à rembourser'}, 
                    status=400
                )
            
            serializer = DebtPaymentSerializer(data=request.data)
            if serializer.is_valid():
                validated_data = serializer.validated_data
                
                # Créer la transaction financière
                from .models import Account, FinancialTransaction
                account = Account.objects.get(id=validated_data['account'])
                
                transaction = FinancialTransaction.objects.create(
                    transaction_type='debt_payment',
                    amount=validated_data['amount'],
                    to_account=account,  # L'argent entre dans ce compte
                    customer=customer,
                    description=validated_data.get('description', f'Remboursement de dette - {customer.name}'),
                    created_by=request.user
                )
                
                # Le signal update_customer_debt_on_payment se chargera de mettre à jour la dette
                
                # Retourner les informations mises à jour
                customer.refresh_from_db()
                return Response({
                    'success': True,
                    'message': f'Remboursement de {validated_data["amount"]} F enregistré',
                    'transaction_number': transaction.transaction_number,
                    'remaining_debt': customer.debt,
                    'account_balance': account.balance
                })
            
            return Response(serializer.errors, status=400)
            
        except Exception as e:
            return Response({'error': f'Erreur lors du remboursement: {str(e)}'}, status=500)


class SupplierViewSet(viewsets.ModelViewSet, StoreContextMixin):
    serializer_class = SupplierSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'phone', 'email', 'address']
    ordering_fields = ['name', 'created_at']
    ordering = ['-created_at']
    
    def get_queryset(self):
        queryset = Supplier.objects.all().order_by('-created_at')
        return self.get_store_queryset(queryset)
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['store'] = self.store
        return context
    
    def create(self, request, *args, **kwargs):
        request.data['store'] = self.store.id
        return super().create(request, *args, **kwargs)
    
    @action(detail=False, methods=['get'], url_path='search')
    def search(self, request):
        """Recherche rapide pour l'autocomplétion"""
        search_query = request.query_params.get('search', '')
        limit = int(request.query_params.get('limit', 20))
        
        if not search_query:
            return Response([])
        
        queryset = self.get_queryset().filter(
            Q(name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(address__icontains=search_query)
        )[:limit]
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class WarehouseViewSet(viewsets.ModelViewSet, StoreContextMixin):
    serializer_class = WarehouseSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'address']
    ordering_fields = ['name', 'created_at']
    ordering = ['-created_at']
    
    def get_queryset(self):
        queryset = Warehouse.objects.all().order_by('-created_at')
        return self.get_store_queryset(queryset)
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['store'] = self.request.user.store
        return context
    
    def create(self, request, *args, **kwargs):
        request.data['store'] = self.store.id
        return super().create(request, *args, **kwargs)
        


class StockEntryViewSet(viewsets.ModelViewSet, StoreContextMixin):
    serializer_class = StockEntrySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['entry_number', 'supplier__name', 'warehouse__name', 'notes']
    ordering_fields = ['created_at', 'total_amount', 'entry_number']
    ordering = ['-created_at']
    
    def get_queryset(self):
        queryset = StockEntry.objects.all().order_by('-created_at')
        return self.get_store_queryset(queryset)
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return StockEntryFormSerializer
        return StockEntrySerializer
    
    def create(self, request, *args, **kwargs):
        from decimal import Decimal
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Créer l'entrée de stock
        supplier = get_object_or_404(Supplier, id=serializer.validated_data['supplier'])
        warehouse = get_object_or_404(Warehouse, id=serializer.validated_data['warehouse'])
        
        with transaction.atomic():
            # Générer un numéro d'entrée unique
            entry_number = f"EN-{StockEntry.objects.count() + 1:06d}"
            
            # Récupérer le compte source si spécifié
            account = None
            account_id = serializer.validated_data.get('account')
           
            if account_id:
                try:
                    account = Account.objects.get(
                        id=account_id, 
                        store=self.request.user.store,
                        is_active=True
                    )
                except Account.DoesNotExist:
                    return Response(
                        {'error': 'Compte spécifié non trouvé ou inactif'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            stock_entry = StockEntry.objects.create(
                entry_number=entry_number,
                supplier=supplier,
                warehouse=warehouse,
                account=account,  # Ajout du compte
                notes=serializer.validated_data.get('notes', ''),
                created_by=self.request.user,
                total_amount=Decimal('0.00')
            )
            
            total_amount = Decimal('0.00')
            
            # Créer les items
            for item_data in serializer.validated_data['items']:
                product = get_object_or_404(Product, id=item_data['product'])
                quantity = int(item_data['quantity'])
                purchase_price = Decimal(item_data['purchase_price'])
                
                StockEntryItem.objects.create(
                    stock_entry=stock_entry,
                    product=product,
                    quantity=quantity,
                    purchase_price=purchase_price,
                    total_price=quantity * purchase_price
                )
                
                total_amount += quantity * purchase_price
            
            # Mettre à jour le montant total
            stock_entry.total_amount = total_amount
            stock_entry.save()
            
            return Response(StockEntrySerializer(stock_entry).data, status=status.HTTP_201_CREATED)


class StockExitViewSet(viewsets.ModelViewSet, StoreContextMixin):
    serializer_class = StockExitSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['exit_number', 'customer__name', 'customer_name', 'warehouse__name', 'notes']
    ordering_fields = ['created_at', 'total_amount', 'exit_number']
    ordering = ['-created_at']
    
    def get_queryset(self):
        queryset = StockExit.objects.all().order_by('-created_at')
        return self.get_store_queryset(queryset)
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return StockExitFormSerializer
        return StockExitSerializer
    
    def create(self, request, *args, **kwargs):
        from decimal import Decimal
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Créer la sortie de stock
        warehouse = get_object_or_404(Warehouse, id=serializer.validated_data['warehouse'])
        customer = None
        if serializer.validated_data.get('customer'):
            customer = get_object_or_404(Customer, id=serializer.validated_data['customer'])
        
        # Validation préalable du stock pour tous les items
        stock_errors = []
        for item_data in serializer.validated_data['items']:
            product = get_object_or_404(Product, id=item_data['product'])
            quantity = int(item_data['quantity'])
            
            try:
                stock = ProductStock.objects.get(
                    product=product,
                    warehouse=warehouse
                )
                if stock.quantity < quantity:
                    stock_errors.append({
                        'product': product.reference,
                        'product_name': product.name,
                        'requested_quantity': quantity,
                        'available_quantity': stock.quantity,
                        'error': 'Stock insuffisant'
                    })
            except ProductStock.DoesNotExist:
                stock_errors.append({
                    'product': product.reference,
                    'product_name': product.name,
                    'requested_quantity': quantity,
                    'available_quantity': 0,
                    'error': 'Aucun stock disponible'
                })
        
        if stock_errors:
            return Response(
                {
                    'error': 'Erreurs de stock détectées',
                    'details': 'Certains produits n\'ont pas suffisamment de stock disponible',
                    'stock_errors': stock_errors,
                    'warehouse': warehouse.name
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            # Générer un numéro de sortie unique
            exit_number = f"EX-{StockExit.objects.count() + 1:06d}"
            
            # Récupérer le compte de destination si spécifié
            account = None
            account_id = serializer.validated_data.get('account')
           
            if account_id:
                print("---------------------------------------------, account1")
                try:
                    account = Account.objects.get(
                        id=account_id, 
                        store=self.request.user.store,
                        is_active=True
                    )
                    print("---------------------------------------------, accoun2", account)
                except Account.DoesNotExist:
                    return Response(
                        {'error': 'Compte spécifié non trouvé ou inactif'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            stock_exit = StockExit.objects.create(
                exit_number=exit_number,
                customer=customer,
                customer_name=serializer.validated_data.get('customer_name', ''),
                warehouse=warehouse,
                account=account,  # Ajout du compte
                notes=serializer.validated_data.get('notes', ''),
                created_by=self.request.user,
                total_amount=Decimal('0.00')
            )
            
            total_amount = Decimal('0.00')
            
            # Créer les items
            for item_data in serializer.validated_data['items']:
                product = get_object_or_404(Product, id=item_data['product'])
                quantity = int(item_data['quantity'])
                sale_price = Decimal(item_data['sale_price'])
                
                try:
                    StockExitItem.objects.create(
                        stock_exit=stock_exit,
                        product=product,
                        quantity=quantity,
                        sale_price=sale_price,
                        total_price=quantity * sale_price
                    )
                except (ValueError, ProductStock.DoesNotExist) as e:
                    # Supprimer le bon de sortie créé en cas d'erreur
                    stock_exit.delete()
                    return Response(
                        {
                            'error': 'Erreur de stock',
                            'details': str(e),
                            'product': product.reference,
                            'product_name': product.name
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
                except ProductStock.DoesNotExist:
                    # Supprimer le bon de sortie créé en cas d'erreur
                    stock_exit.delete()
                    return Response(
                        {
                            'error': 'Stock non disponible',
                            'details': f'Aucun stock disponible pour le produit {product.reference} dans l\'entrepôt {warehouse.name}',
                            'product': product.reference,
                            'warehouse': warehouse.name
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                total_amount += quantity * sale_price
            
            # Mettre à jour le montant total
            stock_exit.total_amount = total_amount
            stock_exit.save()
            
            return Response(StockExitSerializer(stock_exit).data, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def stock_stats(request):
    """Statistiques du stock pour le dashboard"""
    try:
        store = getattr(request.user, 'store', None)
        if not store:
            return Response(
                {'error': 'Store non trouvé'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Calculer les statistiques
        products_count = Product.objects.filter(store=store).count()
        entries_count = StockEntry.objects.filter(
            warehouse__store=store
        ).count()
        exits_count = StockExit.objects.filter(
            warehouse__store=store
        ).count()
        
        # Calculer la valeur totale du stock
        total_stock_value = 0
        products = Product.objects.filter(store=store)
        
        for product in products:
            # Calculer le stock actuel
            entries = StockEntryItem.objects.filter(product=product).aggregate(
                total=Sum('quantity'))['total'] or 0
            exits = StockExitItem.objects.filter(product=product).aggregate(
                total=Sum('quantity'))['total'] or 0
            current_stock = entries - exits
            
            # Obtenir le dernier prix d'achat
            last_entry = StockEntryItem.objects.filter(product=product).order_by('-id').first()
            if last_entry and current_stock > 0:
                total_stock_value += float(last_entry.purchase_price) * current_stock
        
        # Produits en rupture de stock
        low_stock_count = 0
        for product in products:
            entries = StockEntryItem.objects.filter(product=product).aggregate(
                total=Sum('quantity'))['total'] or 0
            exits = StockExitItem.objects.filter(product=product).aggregate(
                total=Sum('quantity'))['total'] or 0
            current_stock = entries - exits
            
            if current_stock <= product.min_stock_alert:
                low_stock_count += 1
        
        return Response({
            'products_count': products_count,
            'entries_count': entries_count,
            'exits_count': exits_count,
            'total_stock_value': total_stock_value,
            'low_stock_count': low_stock_count
        })
        
    except Exception as e:
        logger.error(f"Erreur lors du calcul des statistiques: {str(e)}")
        return Response(
            {'error': 'Erreur interne du serveur'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


class AccountViewSet(viewsets.ModelViewSet, StoreContextMixin):
    """ViewSet pour la gestion des comptes (caisse, banque)"""
    serializer_class = AccountSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name']
    ordering = ['-created_at']
    
    def get_queryset(self):
        return Account.objects.filter(store=self.store)
    
    def perform_create(self, serializer):
        serializer.save(store=self.store)
    
    @action(detail=False, methods=['get'], url_path='active')
    def active_accounts(self, request):
        """Retourne seulement les comptes actifs pour les sélecteurs"""
        queryset = self.get_queryset().filter(is_active=True)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'], url_path='transactions')
    def account_transactions(self, request, pk=None):
        """Retourne l'historique des transactions d'un compte"""
        account = self.get_object()
        from .models import FinancialTransaction
        
        # Récupérer toutes les transactions entrantes et sortantes
        incoming = FinancialTransaction.objects.filter(to_account=account)
        outgoing = FinancialTransaction.objects.filter(from_account=account)
        
        # Combiner et ordonner par date
        all_transactions = incoming.union(outgoing).order_by('-created_at')
        
        # Paginer les résultats
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        
        from django.core.paginator import Paginator
        paginator = Paginator(all_transactions, page_size)
        page_obj = paginator.get_page(page)
        
        # Formater les données
        transactions_data = []
        for transaction in page_obj:
            movement_type = 'credit' if transaction.to_account == account else 'debit'
            transactions_data.append({
                'id': transaction.id,
                'transaction_number': transaction.transaction_number,
                'transaction_type': transaction.transaction_type,
                'transaction_type_display': transaction.get_transaction_type_display(),
                'amount': str(transaction.amount),
                'movement_type': movement_type,
                'description': transaction.description,
                'from_account_name': transaction.from_account.name if transaction.from_account else None,
                'to_account_name': transaction.to_account.name if transaction.to_account else None,
                'stock_exit_number': transaction.stock_exit.exit_number if transaction.stock_exit else None,
                'stock_entry_number': transaction.stock_entry.entry_number if transaction.stock_entry else None,
                'created_by_name': transaction.created_by.fullname,
                'created_at': transaction.created_at.isoformat(),
            })
        
        return Response({
            'count': paginator.count,
            'next': page_obj.has_next(),
            'previous': page_obj.has_previous(),
            'results': transactions_data
        })
    
    def create(self, request, *args, **kwargs):
        """Créer un nouveau compte avec gestion d'erreurs"""
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            if "unique constraint" in str(e).lower():
                return Response(
                    {
                        'error': 'Erreur de création',
                        'details': 'Un compte avec ce nom existe déjà.'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            return Response(
                {
                    'error': 'Erreur lors de la création du compte',
                    'details': str(e)
                },
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def update(self, request, *args, **kwargs):
        """Mettre à jour un compte avec gestion d'erreurs"""
        try:
            return super().update(request, *args, **kwargs)
        except Exception as e:
            if "unique constraint" in str(e).lower():
                return Response(
                    {
                        'error': 'Erreur de modification',
                        'details': 'Un compte avec ce nom existe déjà.'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            return Response(
                {
                    'error': 'Erreur lors de la modification du compte',
                    'details': str(e)
                },
                status=status.HTTP_400_BAD_REQUEST
            )


class InvoiceViewSet(viewsets.ModelViewSet, StoreContextMixin):
    """ViewSet pour la gestion des factures"""
    serializer_class = InvoiceSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['invoice_number', 'customer__name', 'customer_name']
    ordering = ['-created_at']
    
    def get_queryset(self):
        queryset = Invoice.objects.select_related(
            'stock_exit', 'stock_exit__warehouse', 'customer'
        ).prefetch_related('stock_exit__items__product')
        
        # Filtrer par store via le stock_exit -> warehouse -> store
        if self.store:
            queryset = queryset.filter(stock_exit__warehouse__store=self.store)
        else:
            queryset = queryset.none()
            
        return queryset.order_by('-created_at')
    
    @action(detail=True, methods=['get'], url_path='print-data')
    def print_data(self, request, pk=None):
        """Retourne les données formatées pour l'impression de la facture"""
        invoice = self.get_object()
        
        # Données de la boutique
        store_data = {
            'name': invoice.stock_exit.warehouse.store.name,
            'description': invoice.stock_exit.warehouse.store.description,
        }
        
        # Données du client
        customer_data = {
            'name': invoice.customer.name if invoice.customer else invoice.customer_name,
            'phone': invoice.customer.phone if invoice.customer else None,
            'email': invoice.customer.email if invoice.customer else None,
            'address': invoice.customer.address if invoice.customer else None,
        }
        
        # Données des articles
        items_data = []
        for item in invoice.stock_exit.items.all():
            items_data.append({
                'product_reference': item.product.reference,
                'product_name': item.product.name,
                'quantity': item.quantity,
                'unit_price': str(item.sale_price),
                'total_price': str(item.total_price)
            })
        
        # Données complètes de la facture
        invoice_data = {
            'invoice_number': invoice.invoice_number,
            'date': invoice.created_at.strftime('%d/%m/%Y'),
            'time': invoice.created_at.strftime('%H:%M'),
            'total_amount': str(invoice.total_amount),
            'warehouse': invoice.stock_exit.warehouse.name,
            'created_by': invoice.stock_exit.created_by.fullname,
            'store': store_data,
            'customer': customer_data,
            'items': items_data,
            'notes': invoice.stock_exit.notes
        }
        
        return Response(invoice_data)
    
    @action(detail=True, methods=['get'], url_path='download-pdf')
    def download_pdf(self, request, pk=None):
        """Génère et retourne le PDF de la facture"""
        invoice = self.get_object()
        
        # Préparer le contexte pour le template
        context = {
            'invoice': invoice,
            'store': {
                'name': invoice.stock_exit.warehouse.store.name,
                'address': invoice.stock_exit.warehouse.store.description or 'Adresse non renseignée',
                'phone': 'Téléphone non renseigné',  # Vous pouvez ajouter ce champ au modèle Store
                'email': 'Email non renseigné',    # Vous pouvez ajouter ce champ au modèle Store
            },
            'customer': {
                'name': invoice.customer.name if invoice.customer else invoice.customer_name,
                'phone': invoice.customer.phone if invoice.customer else 'Non renseigné',
                'email': invoice.customer.email if invoice.customer else 'Non renseigné',
                'address': invoice.customer.address if invoice.customer else 'Non renseignée',
            },
            'items': []
        }
        
        # Préparer les articles
        for item in invoice.stock_exit.items.all():
            context['items'].append({
                'product': {
                    'name': item.product.name,
                    'barcode': item.product.reference
                },
                'quantity': item.quantity,
                'unit_price': item.sale_price,
                'total_amount': item.total_price
            })
        
        # Calculer les totaux
        subtotal = sum(float(item['total_amount']) for item in context['items'])
        tax_rate = 0.18  # TVA 18%
        tax_amount = subtotal * tax_rate
        total_amount = subtotal + tax_amount
        
        context.update({
            'subtotal': subtotal,
            'tax_amount': tax_amount,
            'total_amount': total_amount,
        })
        
        # Rendre le template HTML
        html_string = render_to_string('invoice_pdf.html', context)
        
        # Générer le PDF avec WeasyPrint
        html = HTML(string=html_string)
        pdf_buffer = io.BytesIO()
        html.write_pdf(pdf_buffer)
        pdf_buffer.seek(0)
        
        # Créer la réponse HTTP avec le PDF
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="facture_{invoice.invoice_number}.pdf"'
        
        return response
    
    def create(self, request, *args, **kwargs):
        return Response(
            {'error': 'Les factures sont créées automatiquement avec les bons de sortie'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )
    
    def update(self, request, *args, **kwargs):
        return Response(
            {'error': 'Les factures ne peuvent pas être modifiées'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )
    
    def destroy(self, request, *args, **kwargs):
        return Response(
            {'error': 'Les factures ne peuvent pas être supprimées'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )


class FinancialTransactionViewSet(viewsets.ModelViewSet, StoreContextMixin):
    """
    ViewSet pour la gestion des transactions financières
    """
    serializer_class = FinancialTransactionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['transaction_type', 'from_account', 'to_account']
    search_fields = ['description', 'transaction_number']
    ordering_fields = ['created_at', 'amount']
    ordering = ['-created_at']

    def get_queryset(self):
        """
        Filtre les transactions par boutique de l'utilisateur connecté
        et ajoute des filtres personnalisés
        """
        user = self.request.user
        
        # Utiliser une seule requête avec Q objects au lieu d'union
        queryset = FinancialTransaction.objects.select_related(
            'from_account', 'to_account', 'created_by', 'stock_entry', 'stock_exit'
        ).filter(
            Q(from_account__store=user.store) | Q(to_account__store=user.store)
        ).distinct().order_by('-created_at')

        # Filtres personnalisés
        account_id = self.request.query_params.get('account_id')
        if account_id:
            queryset = queryset.filter(
                Q(from_account_id=account_id) | Q(to_account_id=account_id)
            )

        date_from = self.request.query_params.get('date_from')
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)

        date_to = self.request.query_params.get('date_to')
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

        return queryset

    def perform_create(self, serializer):
        """
        Associe automatiquement l'utilisateur lors de la création
        """
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Retourne les statistiques des transactions
        """
        queryset = self.get_queryset()
        
        # Calculs des totaux
        sales = queryset.filter(transaction_type='sale')
        purchases = queryset.filter(transaction_type='purchase')
        transfers_in = queryset.filter(transaction_type='transfer_in')
        transfers_out = queryset.filter(transaction_type='transfer_out')
        
        total_sales = sales.aggregate(total=Sum('amount'))['total'] or 0
        total_purchases = purchases.aggregate(total=Sum('amount'))['total'] or 0
        total_transfers_in = transfers_in.aggregate(total=Sum('amount'))['total'] or 0
        total_transfers_out = transfers_out.aggregate(total=Sum('amount'))['total'] or 0
        
        return Response({
            'total_transactions': queryset.count(),
            'total_sales': total_sales,
            'total_purchases': total_purchases,
            'total_transfers_in': total_transfers_in,
            'total_transfers_out': total_transfers_out,
            'net_balance': (total_sales + total_transfers_in) - (total_purchases + total_transfers_out),
            'sales_count': sales.count(),
            'purchases_count': purchases.count(),
            'transfers_in_count': transfers_in.count(),
            'transfers_out_count': transfers_out.count(),
        })

    @action(detail=False, methods=['post'])
    def create_service_payment(self, request):
        """
        Crée une transaction d'apport de service (entrée d'argent dans la caisse)
        """
        from .models import Account, FinancialTransaction
        from decimal import Decimal
        
        # Récupérer les données
        amount = request.data.get('amount')
        description = request.data.get('description', '')
        account_id = request.data.get('account_id')
        
        # Validation des données
        if not amount:
            return Response({'error': 'Le montant est requis'}, status=400)
        
        try:
            amount = Decimal(str(amount))
            if amount <= 0:
                return Response({'error': 'Le montant doit être positif'}, status=400)
        except (ValueError, TypeError):
            return Response({'error': 'Montant invalide'}, status=400)
        
        if not description.strip():
            return Response({'error': 'La description est requise'}, status=400)
        
        # Récupérer le compte de destination
        if account_id:
            try:
                to_account = Account.objects.get(
                    id=account_id,
                    store=self.request.user.store,
                    is_active=True
                )
            except Account.DoesNotExist:
                return Response({'error': 'Compte invalide ou inactif'}, status=400)
        else:
            # Utiliser le premier compte de caisse actif par défaut
            to_account = Account.objects.filter(
                store=self.request.user.store,
                account_type='cash',
                is_active=True
            ).first()
            
            if not to_account:
                # Si pas de caisse, utiliser le premier compte bancaire actif
                to_account = Account.objects.filter(
                    store=self.request.user.store,
                    account_type='bank',
                    is_active=True
                ).first()
            
            if not to_account:
                return Response({'error': 'Aucun compte disponible'}, status=400)
        
        # Créer la transaction
        try:
            transaction = FinancialTransaction.objects.create(
                transaction_type='service',
                amount=amount,
                to_account=to_account,
                description=f"Apport service - {description.strip()}",
                created_by=request.user
            )
            
            # Sérialiser et retourner la transaction créée
            from .serializers import FinancialTransactionSerializer
            serializer = FinancialTransactionSerializer(transaction)
            
            return Response({
                'success': True,
                'message': f'Apport de service enregistré avec succès. Nouveau solde: {to_account.balance} F',
                'transaction': serializer.data
            }, status=201)
            
        except Exception as e:
            return Response({'error': f'Erreur lors de la création: {str(e)}'}, status=500)

    # Export des transactions au format Excel
    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Exporter les transactions en format Excel
        """
        account_id = request.query_params.get('account_id')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        # Récupérer les transactions en fonction des paramètres
        transactions = self.get_queryset()

        if account_id:
            transactions = transactions.filter(
                Q(from_account_id=account_id) | Q(to_account_id=account_id)
            )
        if date_from:
            transactions = transactions.filter(created_at__date__gte=date_from)
        if date_to:
            transactions = transactions.filter(created_at__date__lte=date_to)

        # Créer le fichier Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Ajouter les en-têtes de colonnes
        headers = [
            "ID de Transaction", "Numéro de Transaction", "Montant", "Description", 
            "Type de Transaction", "Compte Source", "Compte Cible", "Créé Par", "Créé Le"
        ]

        ws.append(headers)

        # Ajouter les données des transactions
        for transaction in transactions:
            row = [
                transaction.id,
                transaction.transaction_number,
                transaction.amount,
                transaction.description,
                transaction.get_transaction_type_display(),  # Affichage du type de transaction
                transaction.from_account.name if transaction.from_account else '',
                transaction.to_account.name if transaction.to_account else '',
                transaction.created_by.fullname,
                transaction.created_at.replace(tzinfo=None),
            ]
            ws.append(row)

        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            column = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows():
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Créer une réponse HTTP avec le fichier Excel
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=transactions_export.xlsx"

        # Enregistrer le fichier dans la réponse
        wb.save(response)
        return response

    # Export des transactions au format PDF
    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """
        Exporter les transactions en format PDF
        """
        account_id = request.query_params.get('account_id')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        # Récupérer les transactions en fonction des paramètres
        transactions = self.get_queryset()

        if account_id:
            transactions = transactions.filter(
                Q(from_account_id=account_id) | Q(to_account_id=account_id)
            )
        if date_from:
            transactions = transactions.filter(created_at__date__gte=date_from)
        if date_to:
            transactions = transactions.filter(created_at__date__lte=date_to)

        # Récupérer les informations du store
        store = request.user.store

        # Calculer les statistiques
        total_entrees = transactions.filter(
            transaction_type__in=['purchase', 'service', 'transfer']
        ).aggregate(total=Sum('amount'))['total'] or 0

        total_sorties = transactions.filter(
            transaction_type__in=['sale', 'expense']
        ).aggregate(total=Sum('amount'))['total'] or 0

        solde_net = total_entrees - total_sorties

        # Préparer le contexte pour le template
        context = {
            'transactions': transactions,
            'store': store,
            'total_entrees': total_entrees,
            'total_sorties': total_sorties,
            'solde_net': solde_net,
            'date_from': date_from,
            'date_to': date_to,
            'account_name': '',
        }

        # Si un compte spécifique est sélectionné, récupérer son nom
        if account_id:
            try:
                account = Account.objects.get(id=account_id, store=store)
                context['account_name'] = account.name
            except Account.DoesNotExist:
                pass

        # Rendre le template HTML
        html_string = render_to_string('transactions_pdf.html', context)
        
        # Générer le PDF
        html_doc = HTML(string=html_string)
        pdf_bytes = html_doc.write_pdf()

        # Créer la réponse HTTP
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="rapport_transactions.pdf"'
        
        return response

# 🔄 VIEWSET POUR LES TRANSFERTS DE STOCK
class StockTransferViewSet(viewsets.ModelViewSet):
    """
    ViewSet pour la gestion des transferts de stock
    """
    serializer_class = StockTransferSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = StockTransfer.objects.filter(
            from_warehouse__store=user.store
        ).select_related('from_warehouse', 'to_warehouse', 'created_by').prefetch_related('items__product')
        
        # Filtrage par statut
        status = self.request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filtrage par entrepôt source
        from_warehouse = self.request.query_params.get('from_warehouse')
        if from_warehouse:
            queryset = queryset.filter(from_warehouse_id=from_warehouse)
        
        # Filtrage par entrepôt destination
        to_warehouse = self.request.query_params.get('to_warehouse')
        if to_warehouse:
            queryset = queryset.filter(to_warehouse_id=to_warehouse)
        
        return queryset.order_by('-created_at')
    
    def create(self, request):
        """Créer un nouveau transfert de stock"""
        try:
            # Utiliser le serializer de formulaire pour la validation
            form_serializer = StockTransferFormSerializer(data=request.data)
            if not form_serializer.is_valid():
                return Response(form_serializer.errors, status=400)
            
            validated_data = form_serializer.validated_data
            
            with transaction.atomic():
                # Créer le transfert
                stock_transfer = StockTransfer.objects.create(
                    from_warehouse_id=validated_data['from_warehouse'],
                    to_warehouse_id=validated_data['to_warehouse'],
                    notes=validated_data.get('notes', ''),
                    created_by=request.user
                )
                
                # Traiter les articles
                items = validated_data.get('items', [])
                for item_data in items:
                    product_id = int(item_data['product'])
                    quantity = int(item_data['quantity'])
                    
                    # Vérifier que le produit existe
                    try:
                        product = Product.objects.get(id=product_id)
                    except Product.DoesNotExist:
                        raise ValidationError(f"Produit avec l'ID {product_id} introuvable")
                    
                    # Vérifier le stock disponible dans l'entrepôt source
                    try:
                        source_stock = ProductStock.objects.get(
                            product=product,
                            warehouse_id=validated_data['from_warehouse']
                        )
                        if source_stock.quantity < quantity:
                            raise ValidationError(f"Stock insuffisant pour {product.reference}. Disponible: {source_stock.quantity}, demandé: {quantity}")
                    except ProductStock.DoesNotExist:
                        raise ValidationError(f"Aucun stock disponible pour {product.reference} dans l'entrepôt source")
                    
                    # Créer l'article de transfert
                    StockTransferItem.objects.create(
                        stock_transfer=stock_transfer,
                        product=product,
                        quantity=quantity
                    )
                
                # Sérialiser et retourner le transfert créé
                serializer = self.get_serializer(stock_transfer)
                return Response({
                    'success': True,
                    'message': f'Transfert {stock_transfer.transfer_number} créé avec succès',
                    'transfer': serializer.data
                }, status=201)
                
        except ValidationError as e:
            return Response({'error': str(e)}, status=400)
        except Exception as e:
            return Response({'error': f'Erreur lors de la création: {str(e)}'}, status=500)
    
    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Terminer un transfert (exécuter le mouvement de stock)"""
        try:
            transfer = self.get_object()
            transfer.complete_transfer()
            
            serializer = self.get_serializer(transfer)
            return Response({
                'success': True,
                'message': f'Transfert {transfer.transfer_number} terminé avec succès',
                'transfer': serializer.data
            })
            
        except ValueError as e:
            return Response({'error': str(e)}, status=400)
        except Exception as e:
            return Response({'error': f'Erreur lors de la completion: {str(e)}'}, status=500)
    
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Annuler un transfert"""
        try:
            transfer = self.get_object()
            
            if transfer.status != 'pending':
                return Response({'error': 'Seuls les transferts en attente peuvent être annulés'}, status=400)
            
            transfer.status = 'cancelled'
            transfer.save()
            
            serializer = self.get_serializer(transfer)
            return Response({
                'success': True,
                'message': f'Transfert {transfer.transfer_number} annulé',
                'transfer': serializer.data
            })
            
        except Exception as e:
            return Response({'error': f'Erreur lors de l\'annulation: {str(e)}'}, status=500)